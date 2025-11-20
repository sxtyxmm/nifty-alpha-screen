"""Excel Exporter Module"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List
import logging

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export stock analysis data to Excel format with multiple sheets"""
    
    def __init__(self, output_dir: str = "data/exports"):
        """
        Initialize Excel exporter
        
        Args:
            output_dir: Directory to save Excel files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export(
        self,
        data: pd.DataFrame,
        filename: Optional[str] = None,
        sheet_name: str = "Analysis",
        include_timestamp: bool = True
    ) -> str:
        """
        Export DataFrame to Excel
        
        Args:
            data: DataFrame to export
            filename: Output filename (auto-generated if None)
            sheet_name: Name of the Excel sheet
            include_timestamp: Whether to include timestamp in filename
        
        Returns:
            Path to exported file
        """
        try:
            # Generate filename if not provided
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"stock_analysis_{timestamp}.xlsx" if include_timestamp else "stock_analysis.xlsx"
            
            # Ensure .xlsx extension
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            # Full path
            filepath = self.output_dir / filename
            
            # Export to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                data.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Auto-adjust column widths
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            logger.info(f"Data exported to Excel: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise
    
    def export_multi_sheet(
        self,
        sheets: Dict[str, pd.DataFrame],
        filename: Optional[str] = None,
        include_timestamp: bool = True
    ) -> str:
        """
        Export multiple DataFrames to different sheets in one Excel file
        
        Args:
            sheets: Dictionary of {sheet_name: DataFrame}
            filename: Output filename
            include_timestamp: Whether to include timestamp in filename
        
        Returns:
            Path to exported file
        """
        try:
            # Generate filename
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"stock_analysis_{timestamp}.xlsx" if include_timestamp else "stock_analysis.xlsx"
            
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            filepath = self.output_dir / filename
            
            # Export to Excel with multiple sheets
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                for sheet_name, df in sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Auto-adjust column widths
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            
            logger.info(f"Multi-sheet Excel exported to: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error exporting multi-sheet Excel: {str(e)}")
            raise
    
    def export_with_formatting(
        self,
        data: pd.DataFrame,
        filename: Optional[str] = None,
        conditional_formatting: bool = True
    ) -> str:
        """
        Export with conditional formatting for signals
        
        Args:
            data: DataFrame to export
            filename: Output filename
            conditional_formatting: Whether to apply conditional formatting
        
        Returns:
            Path to exported file
        """
        try:
            from openpyxl.styles import PatternFill
            
            # Export basic file first
            filepath = self.export(data, filename, include_timestamp=True)
            
            if not conditional_formatting or 'Signal' not in data.columns:
                return filepath
            
            # Apply conditional formatting
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Define colors
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            orange_fill = PatternFill(start_color='FFD580', end_color='FFD580', fill_type='solid')
            red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            
            # Find Signal column
            signal_col = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == 'Signal':
                    signal_col = idx
                    break
            
            if signal_col:
                # Apply formatting to each row
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=signal_col)
                    if cell.value == 'BUY':
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = green_fill
                    elif cell.value == 'HOLD':
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = orange_fill
                    elif cell.value == 'AVOID':
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = red_fill
            
            wb.save(filepath)
            logger.info(f"Formatted Excel exported to: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting with formatting: {str(e)}")
            # Return basic export if formatting fails
            return self.export(data, filename)
